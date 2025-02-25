import tkinter as tk
from tkinter import filedialog, messagebox
import asyncio
from googletrans import Translator
import openpyxl
from serbian_text_converter import SerbianTextConverter

# Google Translate language dictionary (same as in original script)
LANGUAGES = {
    "en": "English",
    "sr": "Serbian (Cyrillic)",
    "sr_Latn": "Serbian (Latin)",
    "es": "Spanish",
    "fr": "French",
    "de": "German",
    "it": "Italian",
    "ru": "Russian",
    "pt": "Portuguese",
    "zh-cn": "Chinese (Simplified)",
    "ja": "Japanese",
    "ko": "Korean",
    "ar": "Arabic",
    # Add more languages as necessary
}


class ExcelTranslationApp:
    def __init__(self, input_path, output_path, target_lang="en"):
        self.input_path = input_path
        self.output_path = output_path
        self.target_language_code = target_lang  # Default to "en" (English)

        # Start the translation process
        self.translate_excel_file()

    async def translate_cell(self, cell, translator, target_lang='en'):
        if cell.value and isinstance(cell.value, str):
            original_text = cell.value
            try:
                # Translate the text using Google Translate
                translated_text = await translator.translate(original_text, dest=target_lang)
                print(f"Original: {original_text} -> Translated: {translated_text.text}")

                # If translating to Serbian Latin, convert Cyrillic to Latin
                if target_lang == "sr_Latn":
                    translated_text = SerbianTextConverter.to_latin(translated_text.text)

                # Update the cell value with translated text
                cell.value = translated_text
            except Exception as e:
                print(f"Error translating cell value: {original_text}, Error: {e}")

    async def translate_sheet(self, sheet, translator, target_lang='en'):
        # Iterate through all rows and columns in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                await self.translate_cell(cell, translator, target_lang)

    def translate_excel_file(self):
        if not self.input_path or not self.output_path or not self.target_language_code:
            print("Error: Please provide all required paths and target language.")
            return

        try:
            # Get the current event loop
            loop = asyncio.get_event_loop()

            # Run the async method with the event loop
            loop.run_until_complete(
                self.translate_excel_file_async(self.input_path, self.output_path, self.target_language_code))
        except Exception as e:
            print(f"Error: An error occurred: {e}")

    async def translate_excel_file_async(self, input_path, output_path, target_lang):
        try:
            # Load the Excel file
            workbook = openpyxl.load_workbook(input_path)
            translator = Translator()

            # Process each sheet in the workbook
            for sheet in workbook.sheetnames:
                current_sheet = workbook[sheet]
                print(f"Translating sheet: {sheet}")
                await self.translate_sheet(current_sheet, translator, target_lang)

            # Save the translated workbook
            workbook.save(output_path)
            print(f"Excel file has been translated and saved as {output_path}.")
        except Exception as e:
            print(f"Error: An error occurred: {e}")


class TranslationApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Translator")
        self.root.geometry("700x300")
        self.root.config(bg="#2e2e2e")  # Dark background

        # File path variables
        self.input_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.output_name = tk.StringVar()

        # Target language variables
        self.target_language_code = tk.StringVar()  # Store language code instead of name
        self.target_language_name = tk.StringVar()  # Store the language name for display
        self.target_language_code.set("en")  # Default to English code (en)
        self.target_language_name.set(LANGUAGES["en"])  # Default to "English" as the name

        # Title label
        self.title_label = tk.Label(self.root, text="Excel File Translator", font=("Arial", 16), fg="white", bg="#2e2e2e")
        self.title_label.pack(pady=10)

        # File selection frame
        self.file_frame = tk.Frame(self.root, bg="#2e2e2e")
        self.file_frame.pack(pady=20)

        self.input_file_label = tk.Label(self.file_frame, text="Select Excel file:", font=("Arial", 12), fg="white", bg="#2e2e2e")
        self.input_file_label.pack(side=tk.LEFT, padx=5)

        self.input_file_button = tk.Button(self.file_frame, text="Browse", command=self.select_file, bg="#4a4a4a", fg="white")
        self.input_file_button.pack(side=tk.LEFT)

        self.input_file_entry = tk.Entry(self.file_frame, textvariable=self.input_path, width=40, font=("Arial", 12))
        self.input_file_entry.pack(side=tk.LEFT, padx=5)

        # Language selection frame
        self.language_frame = tk.Frame(self.root, bg="#2e2e2e")
        self.language_frame.pack(pady=10)

        self.language_label = tk.Label(self.language_frame, text="Select target language:", font=("Arial", 12),
                                       fg="white", bg="#2e2e2e")
        self.language_label.pack(side=tk.LEFT, padx=5)

        # Create a list of language names (for display) and their corresponding language codes (for values)
        language_codes = list(LANGUAGES.keys())  # Language codes for values
        language_names = list(LANGUAGES.values())  # Language names for display

        # Function to update language code based on the selected language name
        def update_language_code(*args):
            selected_language_name = self.target_language_name.get()
            for code, name in LANGUAGES.items():
                if name == selected_language_name:
                    self.target_language_code.set(code)
                    break

        # Trace the target_language_name variable so that update_language_code gets called when the selection changes
        self.target_language_name.trace("w", update_language_code)

        # Create the OptionMenu using the language names (displayed) and corresponding language codes (values)
        self.language_dropdown = tk.OptionMenu(self.language_frame, self.target_language_name, *language_names)
        self.language_dropdown.config(font=("Arial", 12), bg="#4a4a4a", fg="white")
        self.language_dropdown.pack(side=tk.LEFT)

        # Output path and filename
        self.output_frame = tk.Frame(self.root, bg="#2e2e2e")
        self.output_frame.pack(pady=10)

        self.output_label = tk.Label(self.output_frame, text="Select output location and name:", font=("Arial", 12), fg="white", bg="#2e2e2e")
        self.output_label.pack(side=tk.LEFT, padx=5)

        self.output_name_entry = tk.Entry(self.output_frame, textvariable=self.output_name, width=40, font=("Arial", 12))
        self.output_name_entry.pack(side=tk.LEFT, padx=5)

        self.output_path_button = tk.Button(self.output_frame, text="Browse", command=self.select_output_location, bg="#4a4a4a", fg="white")
        self.output_path_button.pack(side=tk.LEFT)

        # Translate button
        self.translate_button = tk.Button(self.root, text="Translate", command=self.translate_excel, bg="#4a4a4a", fg="white", font=("Arial", 14))
        self.translate_button.pack(pady=20)

    def select_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.input_path.set(file_path)

            # Automatically update the output path
            output_filename = file_path.split("/")[-1].replace(".xlsx", "") + "-translated.xlsx"
            output_dir = "/".join(file_path.split("/")[:-1])  # Get the directory of the selected file
            output_path = f"{output_dir}/{output_filename}"
            self.output_path.set(output_path)  # Set the output path

            # Ensure the output text box displays the output path correctly
            self.output_name.set(f"{output_dir}/{output_filename}")  # You can also update the filename part if needed

    def select_output_location(self):
        output_dir = filedialog.askdirectory()
        if output_dir:
            # Default to the same directory as input file if not specified
            output_filename = self.input_path.get().split("/")[-1].replace(".xlsx", "") + self.output_name.get()
            self.output_path.set(f"{output_dir}/{output_filename}")

    def translate_excel(self):
        input_path = self.input_path.get()
        output_path = self.output_path.get()
        target_lang = self.target_language_code.get()

        if not input_path or not output_path:
            messagebox.showerror("Input Error", "Please select input and output files.")
            return

        try:
            # Initialize and start the translation app
            app = ExcelTranslationApp(input_path, output_path, target_lang)
            messagebox.showinfo("Success", f"Translation completed and saved as {output_path}")
        except Exception as e:
            messagebox.showerror("Translation Error", f"An error occurred during translation: {e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = TranslationApp(root)
    root.mainloop()
