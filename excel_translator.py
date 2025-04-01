import threading
from openpyxl import load_workbook
from azure.ai.translation.text import TextTranslationClient
from azure.core.credentials import AzureKeyCredential

class ExcelTranslationApp:
    def __init__(self, input_path, output_path, target_lang="en", progress_callback=None):
        self.input_path = input_path
        self.output_path = output_path
        self.target_language_code = target_lang
        self.progress_callback = progress_callback

        # Set up Azure Translator credentials
        self.endpoint = "https://api.cognitive.microsofttranslator.com/"
        key_path = "key.txt"  # The file holding the translator key
        with open(key_path, 'r', encoding='utf-8') as file:
            self.subscription_key = file.read()
        self.region = "westeurope"
        self.client = TextTranslationClient(
            endpoint=self.endpoint,
            credential=AzureKeyCredential(self.subscription_key),
            headers={"Ocp-Apim-Subscription-Region": self.region}
        )

        # Start the translation process
        self.translate_excel_file()

    def translate_cell(self, cell, target_lang='en', progress_counter=None, total_cells=None):
        if cell.value:
            # Check if the value is a string, if not, skip the cell
            if isinstance(cell.value, str):
                original_text = cell.value
                try:
                    response = self.client.translate(
                        body=[original_text], to_language=[target_lang]
                    )
                    translated_text = response[0].translations[0].text
                    print(f"Translated text: {translated_text}")

                    # Update the cell value with translated text
                    cell.value = translated_text
                except Exception as e:
                    print(f"Error translating cell value: {original_text}, Error: {e}")
            else:
                print(f"Skipping non-string cell value: {cell.value} (Type: {type(cell.value)})")

            # Update progress
            if progress_counter is not None and total_cells is not None:
                progress_counter[0] += 1  # Increment the counter
                if self.progress_callback:
                    progress = int((progress_counter[0] / total_cells) * 100)
                    self.progress_callback(progress)

    def process_sheet(self, sheet, target_lang='en'):
        threads = []
        total_cells = sum(1 for row in sheet.iter_rows() for cell in row if isinstance(cell.value, str))
        progress_counter = [0]  # Using a list to allow mutable counter in threads

        # Iterate through all rows and columns in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):  # Only process cells with text
                    # Create a new thread to handle translation
                    thread = threading.Thread(target=self.translate_cell,
                                              args=(cell, target_lang, progress_counter, total_cells))
                    threads.append(thread)
                    thread.start()

        # Wait for all threads to complete
        for thread in threads:
            thread.join()

    def translate_excel_file(self):
        if not self.input_path or not self.output_path or not self.target_language_code:
            print("Error: Please provide all required paths and target language.")
            return

        try:
            # Load the Excel file
            workbook = load_workbook(self.input_path)

            # Process each sheet in the workbook
            for sheet_name in workbook.sheetnames:
                current_sheet = workbook[sheet_name]
                print(f"Translating sheet: {sheet_name}")
                self.process_sheet(current_sheet, self.target_language_code)

            # Save the translated workbook
            workbook.save(self.output_path)
            print(f"Excel file has been translated and saved as {self.output_path}.")
        except Exception as e:
            print(f"Error: An error occurred: {e}")
            if self.progress_callback:
                self.progress_callback(0)  # Reset progress in case of error
